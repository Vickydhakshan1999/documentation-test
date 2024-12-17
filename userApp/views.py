from django_access_point.models.custom_field import CUSTOM_FIELD_STATUS
from django_access_point.models.user import USER_TYPE_CHOICES, USER_STATUS_CHOICES
from django_access_point.views.custom_field import CustomFieldViewSet
from django_access_point.views.crud import CrudViewSet
from django.contrib.auth import get_user_model
from rest_framework.exceptions import PermissionDenied
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth.models import Group
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import permission_classes
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate
from .models import User, UserCustomField, UserCustomFieldValue
from .serializers import UserSerializer, UserCustomFieldSerializer
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
from rest_framework.decorators import api_view, permission_classes
from django.contrib.auth.hashers import check_password
from .serializers import UserDetailsSerializer
import openpyxl
from rest_framework.views import APIView
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from userApp.models import User 
from .permissions import CanCreateUserPermission, CanDeleteUserPermission, CanEditUserPermission, CanViewUserPermission


class PlatformUser(CrudViewSet):
    queryset = User.objects.filter(user_type=USER_TYPE_CHOICES[0][0]).exclude(
        status=USER_STATUS_CHOICES[0][0]
    )
    serializer_class = UserSerializer
    custom_field_model = UserCustomField
    custom_field_value_model = UserCustomFieldValue


    def create(self, request, *args, **kwargs):
        # Initialize the serializer with the request data
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        print(f"Request data: {request.data}") 

        # Save the user without groups initially
        user = serializer.save()

        user.password = request.data.get("password")

        # Get the group ID from the request data
        group_id = request.data.get("group", None)  # Fetch the group ID from request data
        print(f"Group ID from request: {group_id}")  # Debugging line
        

        if group_id:
            # Ensure the group exists and assign it to the user
            try:
                group = Group.objects.get(id=group_id)
                user.groups.clear()  # Clear any existing groups to ensure only the selected group is assigned
                user.groups.add(group)  # Add the group to the user
                user.save()  # Save the user with the assigned group
                print(f"Group '{group.name}' assigned to user '{user.name}'.")  # Debugging line
            except Group.DoesNotExist:
                raise NotFound(detail="Group with the given ID does not exist.")
        
        # Return the response with the saved user data
        return Response(serializer.data, status=status.HTTP_201_CREATED)
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required


  
# Login API

@api_view(['POST'])
def login(request):
    # return/ Response("success")
    email = request.data.get('email')
    password = request.data.get('password')

    if not email or not password:
        return Response({'error': 'Email and password are required.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        # Fetch the user by email
        User = get_user_model()
        user = User.objects.get(email=email)
        # print(user.password)
        # print(password)
        # Compare plain text password directly
        if user.password == password:
            
            # Generate JWT tokens
            refresh = RefreshToken.for_user(user)

            

            return Response({
                "message": "Login successful",
                "user_id": user.id,
                "email": user.email,
                "refresh": str(refresh),
                "access": str(refresh.access_token),
            }, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Incorrect password'}, status=status.HTTP_401_UNAUTHORIZED)
    except User.DoesNotExist:
        return Response({'error': 'User does not exist'}, status=status.HTTP_404_NOT_FOUND)


class PlatformUserCustomField(CustomFieldViewSet):
    queryset = UserCustomField.objects.filter(status=CUSTOM_FIELD_STATUS[1][0])
    serializer_class = UserCustomFieldSerializer


# ....................

# Create a new group
@api_view(['POST'])

def create_group(request):
    """
    This view allows the creation of a new group.
    """
    group_name = request.data.get('name', None)
    if not group_name:
        return Response({'detail': 'Group name is required.'}, status=status.HTTP_400_BAD_REQUEST)

    # Create the group
    group = Group.objects.create(name=group_name)
    
    return Response({'detail': f'Group "{group_name}" created successfully.'}, status=status.HTTP_201_CREATED)


# List all groups
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def list_groups(request):
    """
    This view returns a list of all groups.
    """
    groups = Group.objects.all()
    group_names = [group.name for group in groups]
    return Response({'groups': group_names}, status=status.HTTP_200_OK)


# Retrieve a specific group by its ID
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def retrieve_group(request, group_id):
    """
    This view returns details of a single group.
    """
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        raise NotFound(detail="Group not found")

    return Response({'id': group.id, 'name': group.name}, status=status.HTTP_200_OK)


# Update a group
@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def update_group(request, group_id):
    """
    This view updates the name of an existing group.
    """
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        raise NotFound(detail="Group not found")
    
    new_name = request.data.get('name', None)
    if not new_name:
        return Response({'detail': 'New group name is required.'}, status=status.HTTP_400_BAD_REQUEST)
    
    group.name = new_name
    group.save()

    return Response({'detail': f'Group name updated to "{new_name}".'}, status=status.HTTP_200_OK)


# Delete a group
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_group(request, group_id):
    """
    This view deletes a group.
    """
    try:
        group = Group.objects.get(id=group_id)
    except Group.DoesNotExist:
        raise NotFound(detail="Group not found")
    
    group.delete()

    return Response({'detail': 'Group deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)


# ..................................

# Create a new permission (only for authenticated users)
@api_view(['POST'])
# @permission_classes([IsAuthenticated])
def create_permission(request):
    name = request.data.get('name')
    codename = request.data.get('codename')
    app_label = request.data.get('app_label')
    model_name = request.data.get('model_name')
    
    if not all([name, codename, model_name]):
        return Response(
            {'detail': 'Name, codename, and model_name are required.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        content_type = ContentType.objects.get(app_label=app_label, model=model_name)
        permission = Permission.objects.create(
            name=name,
            codename=codename,
            content_type=content_type
        )
        return Response(
            {'detail': f'Permission "{name}" created successfully.'},
            status=status.HTTP_201_CREATED 
        )
    except ContentType.DoesNotExist:
        return Response(
            {'detail': f'Model "{model_name}" does not exist.'},
            status=status.HTTP_400_BAD_REQUEST
        )


# List all permissions (only for authenticated users)
@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def list_permissions(request):
    permissions = Permission.objects.all().values('id', 'name', 'codename', 'content_type__model')
    return Response({'permissions': list(permissions)}, status=status.HTTP_200_OK)


# Retrieve a specific permission by ID (only for authenticated users)
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def retrieve_permission(request, permission_id):
    try:
        permission = Permission.objects.get(id=permission_id)
        return Response({
            'id': permission.id,
            'name': permission.name,
            'codename': permission.codename,
            'model': permission.content_type.model
        }, status=status.HTTP_200_OK)
    except Permission.DoesNotExist:
        raise NotFound(detail="Permission not found")


# Update a permission by ID (only for authenticated users)
@api_view(['PUT'])
# @permission_classes([IsAuthenticated])
def update_permission(request, permission_id):
    try:
        permission = Permission.objects.get(id=permission_id)
    except Permission.DoesNotExist:
        raise NotFound(detail="Permission not found")
    
    name = request.data.get('name')
    codename = request.data.get('codename')
    
    if name:
        permission.name = name
    if codename:
        permission.codename = codename
    
    permission.save()
    return Response({'detail': 'Permission updated successfully.'}, status=status.HTTP_200_OK)


# Delete a permission by ID (only for authenticated users)
@api_view(['DELETE'])
# @permission_classes([IsAuthenticated])
def delete_permission(request, permission_id):
    try:
        permission = Permission.objects.get(id=permission_id)
        permission.delete()
        return Response({'detail': 'Permission deleted successfully.'}, status=status.HTTP_204_NO_CONTENT)
    except Permission.DoesNotExist:
        raise NotFound(detail="Permission not found")


# ..............
# maping permission to group
# 

@api_view(['POST'])
def assign_permission_to_group(request):
    # Get group_name and permission_codenames from the form data
    group_name = request.data.get('group_name')
    permission_codenames = request.data.getlist('permission_codenames')

    # Validate input
    if not group_name or not permission_codenames:
        return Response(
            {'detail': 'Group name and permission codenames are required.'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if isinstance(permission_codenames[0], str):
        permission_codenames = permission_codenames[0].strip('[]').replace(' ', '').split(',')


    try:
        # Get the group by name
        group = Group.objects.get(name=group_name)
        # print(permission_codenames)

        # Fetch permissions based on the codenames
        permissions = Permission.objects.filter(codename__in=permission_codenames)

        # Validate that all requested permissions exist
        if len(permissions) != len(permission_codenames):
            missing_permissions = set(permission_codenames) - set(permissions.values_list('codename', flat=True))
            return Response(
                {'detail': f'The following permissions do not exist: {", ".join(missing_permissions)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Add permissions to the group
        group.permissions.add(*permissions)

        return Response(
            {'detail': f'Permissions {permission_codenames} added to group "{group_name}" successfully.'},
            status=status.HTTP_200_OK
        )
    except Group.DoesNotExist:
        return Response(
            {'detail': f'Group "{group_name}" does not exist.'},
            status=status.HTTP_400_BAD_REQUEST
        )       
# .........
# mapping user to group
@api_view(['POST'])
# @permission_classes([IsAuthenticated])  # Ensure only authenticated users can use this
def assign_user_to_group(request):
    user_id = request.data.get('user_id')  # User ID to map
    group_name = request.data.get('group_name')  # Group to assign the user to
    
    if not all([user_id, group_name]):
        return Response({'detail': 'user_id and group_name are required.'}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        user = User.objects.get(id=user_id)  # Get the user by ID
        group = Group.objects.get(name=group_name)  # Get the group by name

        # Add the user to the group
        user.groups.add(group)
        return Response({'detail': f'User "{user.phone_no}" assigned to group "{group_name}" successfully.'}, 
                        status=status.HTTP_200_OK)
    
    except User.DoesNotExist:
        return Response({'detail': 'User not found.'}, status=status.HTTP_404_NOT_FOUND)
    
    except Group.DoesNotExist:
        return Response({'detail': 'Group not found.'}, status=status.HTTP_404_NOT_FOUND)





# #  Create API

# @api_view(['POST'])
# @permission_classes([IsAuthenticated])  # Ensures that only authenticated users can call this view
# def create_user_with_token(request):
#     """
#     Create a new user only if the authenticated user has the required permissions.
#     """
#     user = request.user  # Get the authenticated user from the request object
    
#     # Check if the user has permission to create users based on group permissions
#     permission_code = 'add_user_create'  # Replace with the correct permission for your app
#     permission = Permission.objects.filter(codename=permission_code).first()
#     print(permission)
#     if not permission:
#         raise PermissionDenied(f"Permission '{permission_code}' does not exist.")

#     # Check if any group the user belongs to has this permission
#     if not any(group.permissions.filter(id=permission.id).exists() for group in user.groups.all()):
#         raise PermissionDenied("You do not have permission to create users.")

#     # Validate the request data
#     serializer = UserSerializer(data=request.data)
#     serializer.is_valid(raise_exception=True)

#     # Save the new user
#     new_user = serializer.save()

#     # Optionally, assign a group to the new user if provided
#     group_name = request.data.get('group_name')
#     if group_name:
#         try:
#             group = Group.objects.get(name=group_name)
#             new_user.groups.add(group)
#         except Group.DoesNotExist:
#             return Response({"detail": f"Group '{group_name}' does not exist."}, status=status.HTTP_400_BAD_REQUEST)

#     return Response(
#         {
#             "message": "User created successfully.",
#             "user": serializer.data
#         },
#         status=status.HTTP_201_CREATED
#     )



# # Delete Api

# @api_view(['DELETE'])
# @permission_classes([IsAuthenticated])  # Ensures that only authenticated users can call this view
# def delete_user_with_token(request, user_id):
#     """
#     Delete a user only if the authenticated user has the required permissions.
#     """
#     user = request.user  # Get the authenticated user from the request object
    
#     # Check if the user has permission to delete users based on group permissions
#     permission_code = 'can_delete_user'  # Replace with the correct permission for your app
#     permission = Permission.objects.filter(codename=permission_code).first()
#     if not permission:
#         raise PermissionDenied(f"Permission '{permission_code}' does not exist.")

#     # Check if any group the user belongs to has this permission
#     if not any(group.permissions.filter(id=permission.id).exists() for group in user.groups.all()):
#         raise PermissionDenied("You do not have permission to delete users.")

#     # Use get_user_model() to get the correct user model dynamically
#     User = get_user_model()

#     # Check if the user to be deleted exists (using your custom User model)
#     try:
#         user_to_delete = User.objects.get(id=user_id)  # Using the custom User model here
#     except User.DoesNotExist:
#         return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

#     # Delete the user
#     user_to_delete.delete()

#     return Response(
#         {"message": "User deleted successfully."},
#         status=status.HTTP_204_NO_CONTENT
#     )

# # Edit API

# @api_view(['PATCH'])
# @permission_classes([IsAuthenticated])  # Ensures that only authenticated users can call this view
# def edit_user_with_token(request, user_id):
#     """
#     Edit a user only if the authenticated user has the required permissions.
#     """
#     user = request.user  # Get the authenticated user from the request object

#     # Check if the user has permission to edit users based on group permissions
#     permission_code = 'can_edit_user'  # Replace with the correct permission for your app
#     permission = Permission.objects.filter(codename=permission_code).first()
#     if not permission:
#         raise PermissionDenied(f"Permission '{permission_code}' does not exist.")

#     # Check if any group the user belongs to has this permission
#     if not any(group.permissions.filter(id=permission.id).exists() for group in user.groups.all()):
#         raise PermissionDenied("You do not have permission to edit users.")

#     # Use get_user_model() to get the correct user model dynamically
#     User = get_user_model()

#     # Check if the user to be edited exists (using your custom User model)
#     try:
#         user_to_edit = User.objects.get(id=user_id)  # Using the custom User model here
#     except User.DoesNotExist:
#         return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

#     # Ensure that the current user is not trying to edit themselves if restricted (optional)
#     if user.id == user_to_edit.id:
#         return Response({"detail": "You cannot edit your own details."}, status=status.HTTP_400_BAD_REQUEST)

#     # Validate and update the user data using the serializer
#     serializer = UserSerializer(user_to_edit, data=request.data, partial=True)  # partial=True allows updating only the fields sent
#     serializer.is_valid(raise_exception=True)

#     # Save the updated user
#     serializer.save()

#     return Response(
#         {
#             "message": "User updated successfully.",
#             "user": serializer.data
#         },
#         status=status.HTTP_200_OK
#     )



# #  View API
# api_view(['GET'])
# @permission_classes([IsAuthenticated])  # Ensures that only authenticated users can call this view
# def view_user_with_token(request, user_id):
#     """
#     View a user's details only if the authenticated user has the required permissions.
#     """
#     user = request.user  # Get the authenticated user from the request object
    
#     # Check if the user has permission to view users based on group permissions
#     permission_code = 'add_user_view'  
#     permission = Permission.objects.filter(codename=permission_code).first()
#     if not permission:
#         raise PermissionDenied(f"Permission '{permission_code}' does not exist.")

#     # Check if any group the user belongs to has this permission
#     if not any(group.permissions.filter(id=permission.id).exists() for group in user.groups.all()):
#         raise PermissionDenied("You do not have permission to view users.")

#     # Use get_user_model() to get the correct user model dynamically
#     User = get_user_model()

#     # Check if the user to be viewed exists (using your custom User model)
#     try:
#         user_to_view = User.objects.get(id=user_id)  # Using the custom User model here
#     except User.DoesNotExist:
#         return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

#     # Serialize and return the user data
#     serializer = UserSerializer(user_to_view)

#     return Response(
#         {
#             "message": "User details retrieved successfully.",
#             "user": serializer.data
#         },
#         status=status.HTTP_200_OK
#     )






# Create API
@api_view(['POST'])
@permission_classes([CanCreateUserPermission])  
def create_user_with_token(request):
    user = request.user
    
    # Validate the request data
    serializer = UserSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    # Save the new user
    new_user = serializer.save()

    # Optionally, assign a group to the new user if provided
    group_name = request.data.get('group_name')
    if group_name:
        try:
            group = Group.objects.get(name=group_name)
            new_user.groups.add(group)
        except Group.DoesNotExist:
            return Response({"detail": f"Group '{group_name}' does not exist."}, status=status.HTTP_400_BAD_REQUEST)

    return Response(
        {
            "message": "User created successfully.",
            "user": serializer.data
        },
        status=status.HTTP_201_CREATED
    )


# Delete API
@api_view(['DELETE'])
@permission_classes([CanDeleteUserPermission])  
def delete_user_with_token(request, user_id):
    # print("bdhbdba")
    user = request.user
    # print("djbfvbd")

    # Use get_user_model() to get the correct user model dynamically
    User = get_user_model()

    # Check if the user to be deleted exists
    try:
        user_to_delete = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

    # Delete the user
    user_to_delete.delete()

    return Response(
        {"message": "User deleted successfully."},
        status=status.HTTP_204_NO_CONTENT
    )


# Edit API
@api_view(['PATCH'])
@permission_classes([CanEditUserPermission])  
def edit_user_with_token(request, user_id):
    user = request.user

    # Use get_user_model() to get the correct user model dynamically
    User = get_user_model()

    # Check if the user to be edited exists
    try:
        user_to_edit = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

    # Ensure that the current user is not trying to edit themselves if restricted (optional)
    if user.id == user_to_edit.id:
        return Response({"detail": "You cannot edit your own details."}, status=status.HTTP_400_BAD_REQUEST)

    # Validate and update the user data using the serializer
    serializer = UserSerializer(user_to_edit, data=request.data, partial=True)  # partial=True allows updating only the fields sent
    serializer.is_valid(raise_exception=True)

    # Save the updated user
    serializer.save()

    return Response(
        {
            "message": "User updated successfully.",
            "user": serializer.data
        },
        status=status.HTTP_200_OK
    )


# View API
@api_view(['GET'])
@permission_classes([CanViewUserPermission]) 
def view_user_with_token(request, user_id):
    user = request.user

    # Use get_user_model() to get the correct user model dynamically
    User = get_user_model()

    # Check if the user to be viewed exists
    try:
        user_to_view = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

    # Serialize and return the user data
    serializer = UserSerializer(user_to_view)

    return Response(
        {
            "message": "User details retrieved successfully.",
            "user": serializer.data
        },
        status=status.HTTP_200_OK
    )


# User_Detail_Create

@api_view(['POST'])
@permission_classes([IsAuthenticated])  # Ensures that only authenticated users can call this view
def create_user_details(request):
    """
    Create user details only if the authenticated user has the required permissions.
    """
    user = request.user  # Get the authenticated user from the request object
    
    # Check if the user has permission to create user details based on group permissions
    permission_code = 'add_user_details'  
    permission = Permission.objects.filter(codename=permission_code).first()
    
    if not permission:
        raise PermissionDenied(f"Permission '{permission_code}' does not exist.")

    # Check if any group the user belongs to has this permission
    if not any(group.permissions.filter(id=permission.id).exists() for group in user.groups.all()):
        raise PermissionDenied("You do not have permission to create user details.")

    # Validate the request data
    serializer = UserDetailsSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    # Save the new user details
    user_id = request.data.get('user_id')
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

    # Save the user details for the specific user
    user_details = serializer.save(user=user)

    # Optionally, you can perform additional actions like assigning more fields or modifying the details
    # For example, if the address is provided, you could process it further here

    return Response(
        {
            "message": "User details created successfully.",
            "user_details": serializer.data
        },
        status=status.HTTP_201_CREATED
    )


# Export all user to excel

@api_view(['GET'])
def export_users_to_excel(request):
    """
    Export a list of users to an Excel file.
    """
    # Create a workbook and a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Define column headers
    headers = ['ID', 'Username', 'Email', 'Phone Number']

    # Apply bold style to headers
    bold_font = Font(bold=True)

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = bold_font

    # Fetch users data
    users = User.objects.all()

    # Write user data to the sheet
    for row_num, user in enumerate(users, 2):
        
        ws[f'A{row_num}'] = user.id
        ws[f'B{row_num}'] = user.name  
        ws[f'C{row_num}'] = user.email
        ws[f'F{row_num}'] = user.phone_no 

    # Create a response to download the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

# Prepare a list of users with their groups
@api_view(['GET'])
def get_all_users_with_groups(request):
    # Query all users
    users = User.objects.all()

    
    user_data = []
    for user in users:
        user_data.append({
            "id": user.id,
            "username": user.name,
            "email": user.email,
            "groups": [group.name for group in user.groups.all()]  
        })

    return Response(user_data)


#  this code is for ElasticSearch

from elasticsearch_dsl.query import Q
from .documents import UserDocument
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import FormParser, MultiPartParser

class DynamicSearchView(APIView):
    parser_classes = [FormParser, MultiPartParser]

    def post(self, request, *args, **kwargs):
        search_field = request.data.get('field', None)
        search_value = request.data.get('value', None)

        if not search_field or not search_value:
            return Response({"error": "Please provide 'field' and 'value' in the request."}, status=400)

        valid_fields = ['name', 'email', 'phone_no', 'id']
        if search_field not in valid_fields:
            return Response({"error": f"Invalid field '{search_field}'."}, status=400)
        
        print(search_value,"kkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkkk")
        
        # For autocomplete, use 'match_phrase_prefix' query
        if search_field in ['name', 'email', 'phone_no']:
            # user_query = Q("match_phrase_prefix", **{search_field: search_value})
            user_query = Q("wildcard", **{search_field: f"*{search_value}*"})

        elif search_field == 'id':
            # Exact match for IDs
            user_query = Q("term", **{search_field: search_value})
        
        # Perform the search and fetch results
        user_results = UserDocument.search().query(user_query)
        
        # Convert results to a list of dictionaries
        combined_results = {
            'users': [hit.to_dict() for hit in user_results]
        }

        # Handle no results
        if not combined_results['users']:
            return Response({"message": "No users found for the given input."}, status=404)

        return Response(combined_results)
    
